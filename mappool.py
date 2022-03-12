# coding=utf-8
import openpyxl
import json
import re
table=openpyxl.load_workbook('mappool.xlsx')
ws=table['Sheet1']
mappool={}
for i in range(ws.max_row):
    name=ws.cell(row=i+1,column=1).value
    mappool[name]=ws.cell(row=i+1,column=3).value
print(mappool)
with open("mappool.json","w") as f:
     json.dump(mappool,f)
with open("mappool.json", 'r') as load_f:
     load_dict = json.load(load_f)
     print(load_dict)