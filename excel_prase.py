import openpyxl
import json
table=openpyxl.load_workbook('mappool.xlsx')
ws=table['Sheet1']
mappool={}
for i in range(ws.max_row):
    mod=ws.cell(row=i+1,column=1).value
    mod=mod+str(ws.cell(row=i+1,column=2).value)
    mappool[mod]=ws.cell(row=i+1,column=3).value
print(mappool)
with open("mappool.json","w") as f:
     json.dump(mappool,f)
